import uuid
import os
import shutil
from openpyxl import Workbook

from fastapi import FastAPI, File, UploadFile, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from typing import List
from app import omr
from app.email import email
from pydantic import BaseModel , EmailStr



class Temp(BaseModel):
    client_mail : EmailStr


app = FastAPI()

origins = [
    "http://localhost:80","*"
]


app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

async def save_file(path,file):
        tmpImg1 = open(path,'wb')
        tmpImg1.write(file.file.read())
        tmpImg1.close()

def delete_user_dir(user_id):
    shutil.rmtree(user_id) 

def create_sheet(user_id, student_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "STUDENT ID"
    sheet["B1"] = "TEST ID"
    index = 0
    for char in range(67,130):
        Char = chr(char)
        sheet[f"{Char}1"] = "QUESTION {index}"
        index += 1
        if index == 19:
            break
    for row, mark_sheet in enumerate(student_list):
        sheet[f"A{row+2}"] = str(mark_sheet["roll_num"])
        sheet[f"B{row+2}"] = mark_sheet["test_id"]
        index = 0 
        for char in range(67,130):
            Char = chr(char)
            sheet[f"{Char}{row+2}"] = mark_sheet["mark_sheet"][index]
            index += 1
            if index == 19:
                break 
    workbook.save(filename=f"{user_id}/mark_sheet.xlsx")

def email_sheet(path, receiver_email):
    print(type(email))
    message = email.attachSheet(path)
    email.send(receiver_email,message)

def read_data_vertical(arr):
    data = 0
    for y in range(len(arr[0])):
        for x in range(len(arr)):
            if arr[x][y] == 1:
                if x+1 != 1:
                    data = (data*10)+ (x+1)
                else:
                    data = (data*10)+ 0
    return data


def vald_horizontal(sheet_arr, answer_arr):
    for index in range(0,len(answer_arr)):
        if sheet_arr[index] != answer_arr[index]:
            return False
    return True

def calculate_mark(sheet_dict, answer_dict):
    mark_sheet_dict={}
    mark_sheet = []
    
    for key in range(1,21):
        temp = vald_horizontal(sheet_dict[key], answer_dict[key])
        mark_sheet.append(temp)
    
    mark_sheet_dict["roll_num"]   = read_data_vertical(sheet_dict["roll_number"])
    mark_sheet_dict["test_id"]    = read_data_vertical(sheet_dict["test_id"])
    mark_sheet_dict["mark_sheet"] = mark_sheet    
    return mark_sheet_dict



def fetch_dict(path):
    return omr.OMR.main(path)

@app.post("/")
async def sample(file1: UploadFile = File(...),file2: List[UploadFile] = File(...)):
    print(file1.filename)
    for file in file2:
        print(file.filename)
    return {"sucess":"wefdc"}

'''@app.post("/")
async def sample(temp:Temp, answer:UploadFile = File(...) ):
    return {"sucess":temp}'''

@app.post("/validate/")
async def image_fun(client_mail:str = File(...),answerKey: UploadFile = File(...),files: List[UploadFile] = File(...)): 

    user_id = str(uuid.uuid1().int)
    file_count = len(files)
    file_names = list()
    path = user_id+ '/0'
    answer_file = path
    os.mkdir(user_id)
    await save_file(path, answerKey)
    
    for index, file in enumerate(files):
        path = user_id+ '/' +str(index+1)
        file_names.append(path)
        await save_file(path, file)

    answer_sheet = fetch_dict(user_id + '/0')
    student_list =[]
    for index in range(1, file_count+1):
        sheet_dict = fetch_dict(f"{user_id}/{index}")
        student_list.append(calculate_mark(sheet_dict,answer_sheet))
    create_sheet(user_id, student_list)
    email_sheet(f"{user_id}/mark_sheet.xlsx", client_mail)
    delete_user_dir(user_id)
    return {"status":"Excel sheet has been mailed"}